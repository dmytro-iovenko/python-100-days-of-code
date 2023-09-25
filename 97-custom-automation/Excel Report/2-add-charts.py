from openpyxl import load_workbook


# Read workbook and select sheet
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

